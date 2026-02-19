"""
WB Report Bot — Telegram бот для автоматических отчётов Wildberries
Автор: сгенерировано Claude (Anthropic)
"""

import os
import io
import json
import logging
import asyncio
from datetime import datetime, timedelta, date
from collections import defaultdict

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from telegram import Bot, Update
from telegram.ext import (
    Application, CommandHandler, ContextTypes, JobQueue
)
from telegram.constants import ParseMode

# ──────────────────────────────────────────────────────────────
# КОНФИГ (берётся из переменных окружения)
# ──────────────────────────────────────────────────────────────
TELEGRAM_TOKEN = os.environ["TELEGRAM_TOKEN"]   # токен бота от @BotFather
CHAT_ID        = os.environ["CHAT_ID"]          # ваш Telegram ID
WB_API_KEY     = os.environ["WB_API_KEY"]       # API-ключ Wildberries (раздел Статистика)

# Порог для алерта "крупное удержание" (руб)
ALERT_THRESHOLD = float(os.environ.get("ALERT_THRESHOLD", "5000"))

logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(message)s",
    level=logging.INFO,
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────
# WB API
# ──────────────────────────────────────────────────────────────
WB_BASE = "https://statistics-api.wildberries.ru/api/v5/supplier"

def wb_headers():
    return {"Authorization": WB_API_KEY}

def get_report(date_from: str, date_to: str) -> list[dict]:
    """Получить финансовый отчёт за период."""
    url = f"{WB_BASE}/reportDetailByPeriod"
    params = {"dateFrom": date_from, "dateTo": date_to, "limit": 100000}
    resp = requests.get(url, headers=wb_headers(), params=params, timeout=30)
    if resp.status_code == 401:
        raise ValueError("❌ Неверный API-ключ WB. Проверь переменную WB_API_KEY.")
    resp.raise_for_status()
    return resp.json() or []

def get_orders(date_from: str) -> list[dict]:
    """Получить заказы (для отчёта по позициям)."""
    url = f"{WB_BASE}/orders"
    params = {"dateFrom": date_from, "flag": 0}
    try:
        resp = requests.get(url, headers=wb_headers(), params=params, timeout=30)
        resp.raise_for_status()
        return resp.json() or []
    except Exception as e:
        log.warning(f"Не удалось получить заказы: {e}")
        return []

def get_sales(date_from: str) -> list[dict]:
    """Получить продажи."""
    url = f"{WB_BASE}/sales"
    params = {"dateFrom": date_from, "flag": 0}
    try:
        resp = requests.get(url, headers=wb_headers(), params=params, timeout=30)
        resp.raise_for_status()
        return resp.json() or []
    except Exception as e:
        log.warning(f"Не удалось получить продажи: {e}")
        return []

# ──────────────────────────────────────────────────────────────
# АНАЛИТИКА
# ──────────────────────────────────────────────────────────────
def analyze_report(rows: list[dict]) -> dict:
    """Посчитать все метрики из детального отчёта."""
    t = defaultdict(float)
    daily = defaultdict(lambda: defaultdict(float))

    for r in rows:
        doc_type = r.get("doc_type_name", "")
        sale     = r.get("retail_price_withdisc_rub", 0) or 0
        ppvz     = r.get("ppvz_for_pay", 0) or 0          # к перечислению
        delivery = r.get("delivery_rub", 0) or 0
        storage  = r.get("storage_fee", 0) or 0
        penalty  = r.get("penalty", 0) or 0
        deduction= r.get("deduction", 0) or 0
        acceptance=r.get("acceptance", 0) or 0
        rr_dt    = (r.get("rr_dt") or "")[:10]

        if doc_type == "Продажа":
            t["sales_count"] += 1
            t["sales_sum"]   += sale
            t["ppvz_sum"]    += ppvz
            daily[rr_dt]["sales"] += sale
            daily[rr_dt]["ppvz"]  += ppvz
        elif doc_type == "Возврат":
            t["returns_count"] += 1
            t["returns_sum"]   += sale

        t["delivery"]   += delivery
        t["storage"]    += storage
        t["penalty"]    += penalty
        t["deduction"]  += deduction
        t["acceptance"] += acceptance
        daily[rr_dt]["delivery"] += delivery

    # Вознаграждение ВБ = разница перечислено - продажи
    t["wb_commission"] = t["ppvz_sum"] - t["sales_sum"]
    t["total_deductions"] = (
        t["wb_commission"] + t["delivery"] + t["storage"] +
        t["acceptance"] + t["deduction"] + t["penalty"]
    )
    t["net_payout"] = t["ppvz_sum"] - t["delivery"] - t["storage"] - t["acceptance"] - t["deduction"] - t["penalty"]

    # % метрики
    s = t["sales_sum"]
    t["commission_pct"] = (t["wb_commission"] / s * 100) if s else 0
    t["delivery_pct"]   = (t["delivery"]      / s * 100) if s else 0
    t["total_ded_pct"]  = (t["total_deductions"] / s * 100) if s else 0

    return {"totals": dict(t), "daily": dict(daily)}

def analyze_positions(orders: list[dict], sales: list[dict]) -> dict:
    """Анализ по позициям — что продано, что не выкуплено."""
    pos = defaultdict(lambda: {
        "ordered": 0, "sold": 0, "returned": 0,
        "cancelled": 0, "revenue": 0, "name": ""
    })

    for o in orders:
        nm = str(o.get("nmId", "unknown"))
        pos[nm]["name"]    = o.get("subject", "") or o.get("category", "") or nm
        pos[nm]["ordered"] += 1
        if o.get("isCancel"):
            pos[nm]["cancelled"] += 1

    for s in sales:
        nm = str(s.get("nmId", "unknown"))
        pos[nm]["name"]    = s.get("subject", "") or pos[nm]["name"] or nm
        stype = s.get("saleID", "")
        if stype.startswith("S"):
            pos[nm]["sold"]    += 1
            pos[nm]["revenue"] += s.get("priceWithDisc", 0) or 0
        elif stype.startswith("R"):
            pos[nm]["returned"] += 1

    return dict(pos)

def format_compare_message(a1: dict, a2: dict, label1: str, label2: str) -> str:
    """Сравнение двух периодов с % изменения."""
    t1 = a1["totals"]
    t2 = a2["totals"]

    def delta(new, old):
        if old == 0:
            return "➕ новое" if new > 0 else "—"
        pct = (new - old) / abs(old) * 100
        arrow = "📈" if pct > 0 else "📉"
        sign  = "+" if pct > 0 else ""
        return f"{arrow} {sign}{pct:.1f}%"

    def row(label, key, fmt="{:,.0f} ₽"):
        v1 = t1.get(key, 0)
        v2 = t2.get(key, 0)
        d  = delta(v2, v1)
        return f"  {label}\n    {fmt.format(v1)}  →  {fmt.format(v2)}  {d}"

    lines = [
        "🔄 *СРАВНЕНИЕ ПЕРИОДОВ*",
        f"  1️⃣  {label1}",
        f"  2️⃣  {label2}",
        "",
        "💰 *ПРОДАЖИ*",
        row("Выручка",       "sales_sum"),
        row("Кол-во продаж", "sales_count", "{:,.0f} шт."),
        row("Возвраты",      "returns_count", "{:,.0f} шт."),
        "",
        "📉 *УДЕРЖАНИЯ*",
        row("Комиссия ВБ",   "wb_commission"),
        row("Логистика",     "delivery"),
        row("Хранение",      "storage"),
        row("Штрафы",        "penalty"),
        "",
        "✅ *ИТОГО К ПОЛУЧЕНИЮ*",
        row("Чистыми",       "net_payout"),
        "",
    ]

    # Вывод победителя
    n1 = t1.get("net_payout", 0)
    n2 = t2.get("net_payout", 0)
    if n1 and n2:
        if n2 > n1:
            diff = n2 - n1
            lines.append(f"🏆 Период 2️⃣ лучше на *{diff:,.0f} ₽*")
        elif n1 > n2:
            diff = n1 - n2
            lines.append(f"🏆 Период 1️⃣ лучше на *{diff:,.0f} ₽*")
        else:
            lines.append("🤝 Периоды равны по чистой выручке")

    return "\n".join(lines)

def check_alerts(rows: list[dict], threshold: float) -> list[str]:
    """Найти штрафы и крупные удержания."""
    alerts = []
    for r in rows:
        penalty   = r.get("penalty", 0) or 0
        deduction = r.get("deduction", 0) or 0
        dt        = (r.get("rr_dt") or "")[:10]
        nm        = r.get("sa_name") or r.get("nm_id") or ""
        if penalty > 0:
            alerts.append(f"⚠️ Штраф {penalty:,.0f} ₽ за {dt} — товар: {nm}")
        if deduction >= threshold:
            alerts.append(f"🔴 Крупное удержание {deduction:,.0f} ₽ за {dt}")
    return alerts

# ──────────────────────────────────────────────────────────────
# ФОРМИРОВАНИЕ СООБЩЕНИЙ
# ──────────────────────────────────────────────────────────────
def format_weekly_message(analysis: dict, positions: dict, date_from: str, date_to: str, alerts: list[str]) -> str:
    t = analysis["totals"]

    lines = [
        f"📊 *ЕЖЕНЕДЕЛЬНЫЙ ОТЧЁТ ВБ*",
        f"📅 {date_from} — {date_to}",
        "",
        "💰 *ПРОДАЖИ*",
        f"  Заказов/продаж: *{int(t.get('sales_count',0))} шт.*",
        f"  Выручка (розн.): *{t.get('sales_sum',0):,.0f} ₽*",
        f"  Возвратов: {int(t.get('returns_count',0))} шт.",
        "",
        "📉 *УДЕРЖАНИЯ ВБ*",
        f"  Вознаграждение ВБ: {t.get('wb_commission',0):,.0f} ₽ ({t.get('commission_pct',0):.1f}%)",
        f"  Логистика:         {t.get('delivery',0):,.0f} ₽ ({t.get('delivery_pct',0):.1f}%)",
        f"  Хранение:          {t.get('storage',0):,.2f} ₽",
        f"  Приёмка:           {t.get('acceptance',0):,.0f} ₽",
        f"  Прочие удержания:  {t.get('deduction',0):,.0f} ₽",
        f"  Штрафы:            {t.get('penalty',0):,.0f} ₽",
        f"  ─────────────────────────",
        f"  Итого удержано:    *{t.get('total_deductions',0):,.0f} ₽* ({t.get('total_ded_pct',0):.1f}%)",
        "",
        "✅ *ИТОГО К ПОЛУЧЕНИЮ*",
        f"  *{t.get('net_payout',0):,.0f} ₽*",
        "",
    ]

    # Топ позиций
    if positions:
        top = sorted(positions.items(), key=lambda x: -x[1]["revenue"])[:5]
        lines.append("🏆 *ТОП-5 ПОЗИЦИЙ ПО ВЫРУЧКЕ*")
        for i, (nm_id, p) in enumerate(top, 1):
            name = (p["name"] or nm_id)[:30]
            lines.append(
                f"  {i}. {name}\n"
                f"     Продано: {p['sold']} шт. | {p['revenue']:,.0f} ₽"
            )
        lines.append("")

        # Не выкупленные
        not_bought = [(nm, p) for nm, p in positions.items()
                      if p.get("cancelled", 0) > 0 or p.get("returned", 0) > 0]
        if not_bought:
            lines.append("📦 *ОТМЕНЫ И ВОЗВРАТЫ*")
            for nm_id, p in not_bought[:5]:
                name = (p["name"] or nm_id)[:30]
                lines.append(
                    f"  • {name}: отмен {p['cancelled']}, возвратов {p['returned']}"
                )
            lines.append("")

    # Алерты
    if alerts:
        lines.append("🚨 *АЛЕРТЫ*")
        for a in alerts[:10]:
            lines.append(f"  {a}")
        lines.append("")

    lines.append("📎 _Excel-отчёт прикреплён ниже_")
    return "\n".join(lines)

def format_monthly_message(analysis: dict, date_from: str, date_to: str) -> str:
    t = analysis["totals"]
    lines = [
        f"📅 *ЕЖЕМЕСЯЧНЫЙ ИТОГ ВБ*",
        f"🗓 {date_from} — {date_to}",
        "",
        "💰 *ИТОГИ МЕСЯЦА*",
        f"  Продаж: *{int(t.get('sales_count',0))} шт.*",
        f"  Выручка: *{t.get('sales_sum',0):,.0f} ₽*",
        f"  Возвратов: {int(t.get('returns_count',0))} шт.",
        "",
        "📊 *СТРУКТУРА ЗАТРАТ*",
        f"  Комиссия ВБ:    {t.get('wb_commission',0):,.0f} ₽ ({t.get('commission_pct',0):.1f}%)",
        f"  Логистика:      {t.get('delivery',0):,.0f} ₽ ({t.get('delivery_pct',0):.1f}%)",
        f"  Хранение:       {t.get('storage',0):,.2f} ₽",
        f"  Приёмка:        {t.get('acceptance',0):,.0f} ₽",
        f"  Прочие:         {t.get('deduction',0):,.0f} ₽",
        f"  Штрафы:         {t.get('penalty',0):,.0f} ₽",
        f"  ─────────────────",
        f"  Итого удержано: *{t.get('total_deductions',0):,.0f} ₽* ({t.get('total_ded_pct',0):.1f}%)",
        "",
        "💵 *ЧИСТЫМИ НА СЧЁТ*",
        f"  *{t.get('net_payout',0):,.0f} ₽*",
        "",
        "📎 _Подробный Excel-отчёт прикреплён_",
    ]
    return "\n".join(lines)

# ──────────────────────────────────────────────────────────────
# EXCEL ОТЧЁТ
# ──────────────────────────────────────────────────────────────
def make_excel(analysis: dict, positions: dict, label: str) -> bytes:
    """Сгенерировать Excel и вернуть как bytes."""
    BG    = "FF0D0D0D"; HEAD  = "FF111111"; ROW1  = "FF1E1E1E"; ROW2  = "FF242424"
    RED   = "FFDC1E1E"; WHITE = "FFF0F0F0"; GREY  = "FF888888"
    GREEN = "FF27AE60"; YELLOW= "FFFFC107"; ORANGE= "FFFF6B35"

    def fl(c): return PatternFill("solid", fgColor=c)
    def fn(bold=False, sz=10, color=WHITE, italic=False):
        return Font(name="Arial", bold=bold, size=sz, color=color, italic=italic)
    def al(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт"
    ws.sheet_view.showGridLines = False

    col_widths = [3, 32, 18, 14, 3]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r in range(1, 80):
        for c in range(1, 6):
            ws.cell(r, c).fill = fl(BG)

    # Title
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 36
    ws.merge_cells("B2:D2")
    ws["B2"].value = f"  ОТЧЁТ WILDBERRIES — {label}"
    ws["B2"].font = Font(name="Arial", bold=True, size=16, color=RED)
    ws["B2"].fill = fl(BG)

    ws.row_dimensions[3].height = 14
    ws.merge_cells("B3:D3")
    ws["B3"].value = f"  Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["B3"].font = fn(sz=9, color=GREY, italic=True)
    ws["B3"].fill = fl(BG)

    ws.row_dimensions[4].height = 8

    t = analysis["totals"]
    FMT = "#,##0.00 ₽"
    FMT0 = "#,##0 ₽"

    sections = [
        ("ПРОДАЖИ", [
            ("Количество продаж, шт.",      int(t.get("sales_count", 0)),   "#,##0",  WHITE),
            ("Выручка (розн. цена), ₽",     t.get("sales_sum", 0),          FMT0,     WHITE),
            ("К перечислению (итого), ₽",   t.get("ppvz_sum", 0),           FMT0,     WHITE),
            ("Возвратов, шт.",              int(t.get("returns_count", 0)),  "#,##0",  GREY),
        ]),
        ("УДЕРЖАНИЯ WILDBERRIES", [
            ("Вознаграждение ВБ (комиссия), ₽", t.get("wb_commission", 0),    FMT0,   YELLOW),
            ("% от выручки",                    t.get("commission_pct",0)/100, "0.00%",YELLOW),
            ("Логистика (доставка), ₽",         t.get("delivery", 0),          FMT0,   ORANGE),
            ("% от выручки",                    t.get("delivery_pct", 0)/100,  "0.00%",ORANGE),
            ("Хранение на складе, ₽",           t.get("storage", 0),           FMT,    WHITE),
            ("Приёмка товара, ₽",               t.get("acceptance", 0),        FMT0,   WHITE),
            ("Прочие удержания, ₽",             t.get("deduction", 0),         FMT0,   ORANGE),
            ("Штрафы, ₽",                       t.get("penalty", 0),           FMT0,   RED),
            ("ИТОГО УДЕРЖАНО, ₽",               t.get("total_deductions", 0),  FMT0,   RED),
            ("% от выручки",                    t.get("total_ded_pct", 0)/100, "0.00%",RED),
        ]),
        ("ИТОГ", [
            ("Чистыми на счёт, ₽", t.get("net_payout", 0), FMT0, GREEN),
        ]),
    ]

    row_n = 5
    for section_title, items in sections:
        ws.row_dimensions[row_n].height = 22
        ws.merge_cells(f"B{row_n}:D{row_n}")
        ws[f"B{row_n}"].value = f"  {section_title}"
        ws[f"B{row_n}"].font = Font(name="Arial", bold=True, size=11, color=RED)
        ws[f"B{row_n}"].fill = fl(BG)
        row_n += 1

        for i, (label_t, val, fmt, clr) in enumerate(items):
            ws.row_dimensions[row_n].height = 20
            is_total = "ИТОГО" in label_t or label_t.startswith("Чистыми")
            bg = HEAD if is_total else (ROW1 if i % 2 == 0 else ROW2)

            ws.merge_cells(f"B{row_n}:C{row_n}")
            ws[f"B{row_n}"].value = label_t
            ws[f"B{row_n}"].font = fn(bold=is_total, sz=10, color=clr)
            ws[f"B{row_n}"].fill = fl(bg)

            cell = ws.cell(row_n, 4)
            cell.value = val
            cell.font = fn(bold=is_total, sz=11 if is_total else 10, color=clr)
            cell.alignment = al("right")
            cell.fill = fl(bg)
            cell.number_format = fmt
            ws.cell(row_n, 5).fill = fl(bg)
            row_n += 1

        row_n += 1  # spacer

    # Positions sheet
    if positions:
        ws2 = wb.create_sheet("Позиции")
        ws2.sheet_view.showGridLines = False
        col_w2 = [3, 30, 10, 10, 10, 10, 14, 3]
        for i, w in enumerate(col_w2, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        for r in range(1, 100):
            for c in range(1, 9):
                ws2.cell(r, c).fill = fl(BG)

        ws2.row_dimensions[1].height = 8
        ws2.row_dimensions[2].height = 30
        ws2.merge_cells("B2:G2")
        ws2["B2"].value = "  АНАЛИЗ ПО ПОЗИЦИЯМ"
        ws2["B2"].font = Font(name="Arial", bold=True, size=14, color=RED)
        ws2["B2"].fill = fl(BG)

        ws2.row_dimensions[3].height = 8
        ws2.row_dimensions[4].height = 26
        hdrs = ["", "Товар", "Заказ.", "Продано", "Отменено", "Возврат", "Выручка, ₽", ""]
        for col, h in enumerate(hdrs, 1):
            c = ws2.cell(4, col)
            c.value = h; c.font = fn(True, 9, GREY)
            c.alignment = al("center", "center", True); c.fill = fl(HEAD)

        sorted_pos = sorted(positions.items(), key=lambda x: -x[1]["revenue"])
        for i, (nm_id, p) in enumerate(sorted_pos):
            rn = 5 + i
            ws2.row_dimensions[rn].height = 18
            bg = ROW1 if i % 2 == 0 else ROW2
            row_v = ["", p["name"] or nm_id, p["ordered"], p["sold"],
                     p["cancelled"], p["returned"], p["revenue"], ""]
            row_f = [None, None, "#,##0", "#,##0", "#,##0", "#,##0", FMT0, None]
            row_c = [WHITE, WHITE, WHITE, GREEN if p["sold"] > 0 else GREY,
                     ORANGE if p["cancelled"] > 0 else GREY,
                     RED if p["returned"] > 0 else GREY, WHITE, WHITE]
            for col, (val, fmt, clr) in enumerate(zip(row_v, row_f, row_c), 1):
                cell = ws2.cell(rn, col)
                cell.value = val; cell.font = fn(sz=9.5, color=clr)
                cell.alignment = al("right" if col >= 3 else "left")
                cell.fill = fl(bg)
                if fmt and val != "": cell.number_format = fmt

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ──────────────────────────────────────────────────────────────
# JOB FUNCTIONS
# ──────────────────────────────────────────────────────────────
async def send_report(bot: Bot, period_label: str, date_from: str, date_to: str):
    """Получить данные с ВБ и отправить отчёт."""
    log.info(f"Формирую отчёт: {date_from} — {date_to}")
    await bot.send_message(
        chat_id=CHAT_ID,
        text=f"⏳ Формирую {period_label} отчёт WB за {date_from} — {date_to}...",
    )
    try:
        rows   = get_report(date_from, date_to)
        orders = get_orders(date_from)
        sales_data = get_sales(date_from)

        if not rows:
            await bot.send_message(
                chat_id=CHAT_ID,
                text=f"ℹ️ Нет данных за период {date_from} — {date_to}. "
                     "Возможно, данные ещё не обновились — попробую позже.",
            )
            return

        analysis  = analyze_report(rows)
        positions = analyze_positions(orders, sales_data)
        alerts    = check_alerts(rows, ALERT_THRESHOLD)

        # Текстовый отчёт
        if "месяц" in period_label.lower():
            msg = format_monthly_message(analysis, date_from, date_to)
        else:
            msg = format_weekly_message(analysis, positions, date_from, date_to, alerts)

        await bot.send_message(
            chat_id=CHAT_ID,
            text=msg,
            parse_mode=ParseMode.MARKDOWN,
        )

        # Алерты отдельным сообщением если есть
        if alerts:
            alert_text = "🚨 *ВНИМАНИЕ — ШТРАФЫ И УДЕРЖАНИЯ*\n\n" + "\n".join(alerts)
            await bot.send_message(
                chat_id=CHAT_ID,
                text=alert_text,
                parse_mode=ParseMode.MARKDOWN,
            )

        # Excel
        excel_bytes = make_excel(analysis, positions, f"{period_label} {date_from}—{date_to}")
        filename    = f"WB_{'weekly' if 'недел' in period_label.lower() else 'monthly'}_{date_from}_{date_to}.xlsx"
        await bot.send_document(
            chat_id=CHAT_ID,
            document=io.BytesIO(excel_bytes),
            filename=filename,
            caption=f"📊 Excel-отчёт за {date_from} — {date_to}",
        )
        log.info("Отчёт успешно отправлен.")

    except Exception as e:
        log.error(f"Ошибка при формировании отчёта: {e}", exc_info=True)
        await bot.send_message(
            chat_id=CHAT_ID,
            text=f"❌ Ошибка при получении данных с WB:\n`{e}`\n\n"
                 "Проверь API-ключ и повтори /report",
            parse_mode=ParseMode.MARKDOWN,
        )

async def weekly_job(context: ContextTypes.DEFAULT_TYPE):
    """Еженедельный отчёт — каждый понедельник за прошлую неделю."""
    today    = date.today()
    date_to  = today - timedelta(days=today.weekday() + 1)   # воскресенье
    date_from= date_to - timedelta(days=6)                   # понедельник
    await send_report(
        context.bot,
        "Еженедельный",
        date_from.strftime("%Y-%m-%d"),
        date_to.strftime("%Y-%m-%d"),
    )

async def monthly_job(context: ContextTypes.DEFAULT_TYPE):
    """Ежемесячный отчёт — 1-го числа за прошлый месяц."""
    today    = date.today()
    first_day= today.replace(day=1)
    last_month_end   = first_day - timedelta(days=1)
    last_month_start = last_month_end.replace(day=1)
    await send_report(
        context.bot,
        "Ежемесячный",
        last_month_start.strftime("%Y-%m-%d"),
        last_month_end.strftime("%Y-%m-%d"),
    )

# ──────────────────────────────────────────────────────────────
# КОМАНДЫ
# ──────────────────────────────────────────────────────────────
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 Привет! Я бот для отчётов Wildberries.\n\n"
        "📌 Команды:\n"
        "/report — отчёт за последние 7 дней\n"
        "/week — отчёт за прошлую неделю\n"
        "/month — отчёт за прошлый месяц\n"
        "/today — отчёт за сегодня\n"
        "/period ДД.ММ.ГГГГ ДД.ММ.ГГГГ — отчёт за любой период\n"
        "/compare Д1 Д1 Д2 Д2 — сравнить два периода\n"
        "/status — проверить подключение к WB\n\n"
        "⏰ Автоотчёты:\n"
        "  • Еженедельно — каждый понедельник в 09:00\n"
        "  • Ежемесячно — 1-го числа в 09:00",
    )

async def cmd_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("🔄 Проверяю подключение к WB API...")
    try:
        today = date.today()
        rows = get_report(today.strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d"))
        await update.message.reply_text(
            f"✅ Подключение работает!\n"
            f"Получено строк за сегодня: {len(rows)}\n"
            f"API-ключ: {'*' * 20}{WB_API_KEY[-4:]}",
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка подключения:\n{e}")

async def cmd_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отчёт за последние 7 дней."""
    today     = date.today()
    date_from = (today - timedelta(days=7)).strftime("%Y-%m-%d")
    date_to   = today.strftime("%Y-%m-%d")
    await send_report(context.bot, "Еженедельный", date_from, date_to)

async def cmd_week(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отчёт за прошлую неделю."""
    today     = date.today()
    date_to   = today - timedelta(days=today.weekday() + 1)
    date_from = date_to - timedelta(days=6)
    await send_report(
        context.bot, "Еженедельный",
        date_from.strftime("%Y-%m-%d"),
        date_to.strftime("%Y-%m-%d"),
    )

async def cmd_month(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отчёт за прошлый месяц."""
    today    = date.today()
    first_day= today.replace(day=1)
    end      = first_day - timedelta(days=1)
    start    = end.replace(day=1)
    await send_report(
        context.bot, "Ежемесячный",
        start.strftime("%Y-%m-%d"),
        end.strftime("%Y-%m-%d"),
    )

async def cmd_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отчёт за сегодня."""
    today = date.today().strftime("%Y-%m-%d")
    await send_report(context.bot, "Дневной", today, today)

async def cmd_period(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отчёт за произвольный период. Использование: /period ДД.ММ.ГГГГ ДД.ММ.ГГГГ"""
    if len(context.args) != 2:
        await update.message.reply_text(
            "📅 Укажи две даты через пробел:\n"
            "`/period ДД.ММ.ГГГГ ДД.ММ.ГГГГ`\n\n"
            "Примеры:\n"
            "`/period 01.01.2025 31.01.2025`\n"
            "`/period 10.02.2025 19.02.2025`",
            parse_mode=ParseMode.MARKDOWN,
        )
        return

    try:
        date_from = datetime.strptime(context.args[0], "%d.%m.%Y").date()
        date_to   = datetime.strptime(context.args[1], "%d.%m.%Y").date()
    except ValueError:
        await update.message.reply_text(
            "❌ Неверный формат даты. Используй ДД.ММ.ГГГГ\n"
            "Пример: `/period 01.02.2025 28.02.2025`",
            parse_mode=ParseMode.MARKDOWN,
        )
        return

    if date_from > date_to:
        await update.message.reply_text(
            "❌ Начальная дата не может быть позже конечной."
        )
        return

    if (date_to - date_from).days > 365:
        await update.message.reply_text(
            "❌ Период не может быть больше 365 дней."
        )
        return

    await send_report(
        context.bot,
        "Произвольный",
        date_from.strftime("%Y-%m-%d"),
        date_to.strftime("%Y-%m-%d"),
    )

async def cmd_compare(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Сравнение двух периодов.
    Использование: /compare ДД.ММ.ГГГГ ДД.ММ.ГГГГ ДД.ММ.ГГГГ ДД.ММ.ГГГГ
    Пример: /compare 01.01.2025 31.01.2025 01.02.2025 28.02.2025
    """
    HELP = (
        "📊 Сравни два периода:\n"
        "`/compare ДД.ММ.ГГГГ ДД.ММ.ГГГГ ДД.ММ.ГГГГ ДД.ММ.ГГГГ`\n\n"
        "Первые две даты — период 1️⃣, вторые две — период 2️⃣\n\n"
        "Примеры:\n"
        "`/compare 01.01.2025 31.01.2025 01.02.2025 28.02.2025`\n"
        "`/compare 01.02.2025 07.02.2025 08.02.2025 14.02.2025`"
    )

    if len(context.args) != 4:
        await update.message.reply_text(HELP, parse_mode=ParseMode.MARKDOWN)
        return

    try:
        d1_from = datetime.strptime(context.args[0], "%d.%m.%Y").date()
        d1_to   = datetime.strptime(context.args[1], "%d.%m.%Y").date()
        d2_from = datetime.strptime(context.args[2], "%d.%m.%Y").date()
        d2_to   = datetime.strptime(context.args[3], "%d.%m.%Y").date()
    except ValueError:
        await update.message.reply_text(
            "❌ Неверный формат даты. Используй ДД.ММ.ГГГГ\n\n" + HELP,
            parse_mode=ParseMode.MARKDOWN,
        )
        return

    if d1_from > d1_to or d2_from > d2_to:
        await update.message.reply_text("❌ Начальная дата не может быть позже конечной.")
        return

    label1 = f"{context.args[0]} — {context.args[1]}"
    label2 = f"{context.args[2]} — {context.args[3]}"

    await update.message.reply_text(
        f"⏳ Загружаю данные для сравнения...\n1️⃣ {label1}\n2️⃣ {label2}"
    )

    try:
        rows1 = get_report(d1_from.strftime("%Y-%m-%d"), d1_to.strftime("%Y-%m-%d"))
        rows2 = get_report(d2_from.strftime("%Y-%m-%d"), d2_to.strftime("%Y-%m-%d"))

        if not rows1 and not rows2:
            await update.message.reply_text("ℹ️ Нет данных ни за один из периодов.")
            return

        a1 = analyze_report(rows1) if rows1 else {"totals": {}, "daily": {}}
        a2 = analyze_report(rows2) if rows2 else {"totals": {}, "daily": {}}

        msg = format_compare_message(a1, a2, label1, label2)
        await update.message.reply_text(msg, parse_mode=ParseMode.MARKDOWN)

    except Exception as e:
        log.error(f"Ошибка /compare: {e}", exc_info=True)
        await update.message.reply_text(
            f"❌ Ошибка при получении данных:\n`{e}`",
            parse_mode=ParseMode.MARKDOWN,
        )

# ──────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────
def main():
    app = Application.builder().token(TELEGRAM_TOKEN).build()

    # Команды
    app.add_handler(CommandHandler("start",  cmd_start))
    app.add_handler(CommandHandler("status", cmd_status))
    app.add_handler(CommandHandler("report", cmd_report))
    app.add_handler(CommandHandler("week",   cmd_week))
    app.add_handler(CommandHandler("month",  cmd_month))
    app.add_handler(CommandHandler("today",  cmd_today))
    app.add_handler(CommandHandler("period", cmd_period))
    app.add_handler(CommandHandler("compare",cmd_compare))

    # Расписание (время — МСК UTC+3)
    jq = app.job_queue
    # Каждый понедельник в 09:00 МСК
    jq.run_daily(weekly_job,  time=datetime.strptime("06:00", "%H:%M").time(), days=(0,))
    # 1-го числа каждого месяца в 09:00 МСК
    jq.run_monthly(monthly_job, when=datetime.strptime("06:00", "%H:%M").time(), day=1)

    log.info("Бот запущен. Ожидаю команды...")
    app.run_polling(allowed_updates=["message"])

if __name__ == "__main__":
    main()
